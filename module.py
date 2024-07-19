from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import jieba
from html import unescape
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from openpyxl import Workbook
from difflib import SequenceMatcher
import allure

class miaoxiang: 
    #初始化依次填写精确度、轮次、文件名、模块xpath表达式
    
    @allure.step("初始化")
    def __init__(self,ROUND=1,url='',answer_file_name="",module_xpath_expression="",question_file_name=""):
        self.url = url
        self.module_xpath_expression = module_xpath_expression
        self.ROUND = ROUND
        self.i = 0
        self.answer_file_name = answer_file_name
        self.question_file_name = question_file_name
        self.chrome_options = Options()
        #填写chrome浏览器userdata存储位置
        self.chrome_options.add_argument(r'--user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data')
        #用于表格内输入问题
        self.row = 1
        
    @allure.step("打开浏览器")
    def test_openwindow(self):
        self.driver = webdriver.Chrome(options=self.chrome_options)
        #进入妙想主页
        time.sleep(0.5)
        self.driver.get(str(self.url))
        time.sleep(0.5)
        self.driver.maximize_window()
        #先使用sleep 暴力等待
        time.sleep(1)

    @allure.step("按键存在断言")
    #按键存在断言
    def test_assert_exist_key(self,key):
        assert key,"按键不存在！！！"
        key.click()
        time.sleep(1)

    @allure.step("提取表格内提问")
    def test_load_xlsx_question(self):
        assert os.path.exists(self.question_file_name),"提问表格文件不存在!!!"

        wb = Workbook()
        wb = load_workbook(self.question_file_name)
        ws = wb.active if wb.active else wb.create_sheet()
        max_row = ws.max_row
        assert self.row <= max_row, "已遍历至提问表格末尾，结束提问!!!"
        question = ws[f'A{self.row}']
        self.row = self.row + 1
        return question
    
    @allure.step("问题输入网页placeholder")
    def test_input_question_in_placeholder(self):
        question = self.test_load_xlsx_question()
        #问题输入框
        questionbox_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/footer/main/textarea'
        #通过xpath寻找问题输入框
        questionbox = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, questionbox_xpath_expression)))
        self.test_assert_exist_key(questionbox)
        questionbox.send_keys(question.value)

    #回答写入表格
    @allure.step("回答写入表格")
    def test_write_answer(self):
        if not os.path.exists(self.answer_file_name):
            print("表格文件不存在")
            wb = Workbook()
            wb.save(self.answer_file_name)
            print(f"文件 {self.answer_file_name} 已创建。")
        wb = load_workbook(self.answer_file_name)
        ws = wb.active if wb.active else wb.create_sheet()

        column_letter = 'C'     #更新前模型使用B 更新后使用C
        max_row = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=ord(column_letter) - ord('A') + 1)
            if cell.value is not None:
                 max_row = row             #获取该列深度

        ws[f'A{max_row + 1}'] = placeholder_text
        ws[f'C{max_row + 1}'] = text_answer  #更新前模型使用B 更新后使用C

        wb.save(self.answer_file_name)
        print("问答存入表格成功")


    #点击侧边栏
    @allure.step("点击侧边栏")
    def test_touch_expand(self):
        #xpath 判断侧边栏是否拉开 未拉开则需要拉开
        #此处xpath对应展开键
        expandkey_xpath_expression = '//*[@id="root"]/div/article/aside[1]/aside'

        #通过xpath寻找拉开键
        expandkey = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, expandkey_xpath_expression)))
        #expandkey = self.driver.find_element(By.XPATH, expandkey_xpath_expression)

        self.test_assert_exist_key(expandkey)

    #点击妙想助手
    @allure.step("点击妙想助手")
    def test_touch_helper(self):
        #妙想助手按钮xpath
        helperkey_xpath_expression = '//*[@id="root"]/div/article/aside[1]/div/header/button[2]'

        #通过xpath寻找妙想助手按钮
        helperkey = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, helperkey_xpath_expression)))
        #helperkey = self.driver.find_element(By.XPATH, helperkey_xpath_expression)

        self.test_assert_exist_key(helperkey)

    #点击对应模块按钮
    @allure.step("点击测试模块")
    def test_touch_module(self,module_key):
        
        self.test_assert_exist_key(module_key)

    #点击问题输入框
    @allure.step("点击问题输入框")
    def test_touch_questionbox(self):
        #问题输入框
        questionbox_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/footer/main/textarea'
        #通过xpath寻找问题输入框
        questionbox = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, questionbox_xpath_expression)))
        #questionbox = self.driver.find_element(By.XPATH, questionbox_xpath_expression)

        self.test_assert_exist_key(questionbox)

    #点击随机推荐问题
    @allure.step("点击随机推荐问题")
    def test_touch_randomquestion(self):
        #推荐问题n按钮xpath
        n = random.randint(1,3)
        recommendquestionkey_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/main/div[1]/article/main/div/div/main/article/div['+str(n)+']'

        #通过xpath寻找推荐问题n按钮
        recommendquestionkey = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, recommendquestionkey_xpath_expression)))
        #recommendquestionkey = self.driver.find_element(By.XPATH, recommendquestionkey_xpath_expression)

        self.test_assert_exist_key(recommendquestionkey)

    #点击发送按键
    @allure.step("点击发送按钮")
    def test_touch_send(self):
        #发送按钮xpath
        sendkey_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/footer/main/section[2]'

        #通过xpath寻找发送按钮
        sendkey = WebDriverWait(self.driver, 3, 0.5).until(EC.presence_of_element_located((By.XPATH, sendkey_xpath_expression)))
        #sendkey = self.driver.find_element(By.XPATH, sendkey_xpath_expression)
        self.test_assert_exist_key(sendkey)
        #暂未有更好等待策略
        time.sleep(40)

    #对比模型改进前后回答差异
    @allure.step("导入回答表格对比前后差异")
    def test_compare_answer(self):
        assert os.path.exists(self.answer_file_name),"待比较回答表格文件不存在!!!"

        wb = load_workbook(self.answer_file_name)
        ws = wb.active
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=2,max_row=max_row):
            text1 = str(row[1].value)
            text2 = str(row[2].value)
            similarity_score = SequenceMatcher(None, text1, text2).ratio()
            ws.cell(row=row[0].row, column=4, value=similarity_score)

        wb.save(self.answer_file_name)


    #爬取输入框问题
    @allure.step("爬取输入框问题")
    def test_get_placeholder_text(self):
        global placeholder_text
        placeholder_text_class_name = "ChatRoom-module__input___BotWC"
        placeholder_text = self.driver.find_element(By.CLASS_NAME,placeholder_text_class_name)
        placeholder_text = placeholder_text.text
        placeholder_text = unescape(placeholder_text)
        #print(placeholder_text.text)
        #time.sleep(1)

    #爬取前端文字回答
    @allure.step("爬取前端回答")
    def test_get_text_answer(self):
        global text_answer
        text_answer_class_name = "ShowType40-module__content___CWmch"
        text_answer = self.driver.find_element(By.CLASS_NAME,text_answer_class_name)
        text_answer = text_answer.text
        text_answer = unescape(text_answer)


    #比对问题回答的相似度
    @allure.step("比对问题回答的相似度")
    def test_compare_similarity(self):
        numofwords_in_placeholder = 0
        numofwords_in_text_answer = 0
        words_in_placeholder = list(jieba.cut(placeholder_text,cut_all=False))
        words_in_text_answer = list(jieba.cut(text_answer,cut_all=False))
        numofwords_in_placeholder = sum(1 for _ in words_in_placeholder)
        #numofwords_in_text_answer = sum(1 for _ in words_in_text_answer)
        set_words_in_placeholder = set(words_in_placeholder)
        set_words_in_text_answer = set(words_in_text_answer)
        
        common_words = set_words_in_placeholder & set_words_in_text_answer

        Accuracy = float(len(common_words)) / float(numofwords_in_placeholder)
        self.test_assert_accuracy(Accuracy)

    #测试循环（页面随机问题）
    @allure.step("进入测试循环1")
    def test_round_1(self):
        self.test_openwindow()
        self.test_touch_expand()
        self.test_touch_helper()
        
        while self.i < self.ROUND:

            #通过xpath寻找模块按钮
            module_key = self.driver.find_element(By.XPATH, self.module_xpath_expression)
            
            self.test_touch_module(module_key)

            self.test_touch_randomquestion()

            self.test_get_placeholder_text()
            
            self.test_touch_send()

            self.test_get_text_answer()

            self.test_compare_similarity()

            self.test_touch_expand()

            self.test_touch_helper()

            self.i = self.i + 1

        self.driver.quit()

    #测试循环（表格导入问题）
    @allure.step("进入测试循环2")
    def test_round_2(self):
        self.test_openwindow()
        self.test_touch_expand()
        self.test_touch_helper()    
        
        while self.i < self.ROUND:

            #通过xpath寻找模块按钮
            module_key = self.driver.find_element(By.XPATH, self.module_xpath_expression)
            
            self.test_touch_module(module_key)

            self.test_input_question_in_placeholder()

            self.test_get_placeholder_text()
            
            self.test_touch_send()

            self.test_get_text_answer()

            self.test_write_answer()

            self.test_touch_expand()

            self.test_touch_helper()

            self.i = self.i + 1

        self.driver.quit()

    #测试文本对比相似度
    @allure.step("进入测试3")
    def test_round_3(self):
        self.test_compare_answer()