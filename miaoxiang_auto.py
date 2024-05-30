from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import random
import jieba
from html import unescape
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from openpyxl import Workbook
import difflib


##################################################初始化测试板块配置##################################################
module_xpath_expression = ['//*[@id="root"]/div/article/main/div/main/article/section[1]/main/div[1]/div',
                           '//*[@id="root"]/div/article/main/div/main/article/section[1]/main/div[2]/div',
                           '//*[@id="root"]/div/article/main/div/main/article/section[2]/main/div[2]/div',
                           '//*[@id="root"]/div/article/main/div/main/article/section[2]/main/div[3]/div']
#精确度阈值
ACCURACY = float(input("输入精确度阈值： "))
#测试轮次
ROUND = int(input("输入测试轮次数： "))
i = 0
#文件名
file_name = "test.xlsx"
chrome_options = Options()
#使用本地浏览器缓存 跳过登录
chrome_options.add_argument(r'--user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data')
#导入本地浏览器配置
driver = webdriver.Chrome(options=chrome_options)
#####################################################################################################################


#按键存在断言
def assert_exist_key(key):
    assert key,"按键不存在！！！"
    key.click()
    time.sleep(1)

#结果精确度断言
def assert_accuracy(Accuracy):
    if Accuracy > ACCURACY:
        print("准确度合格：",Accuracy)
    else:
        print("准确度不合格，问答存入表格：",Accuracy)
        #问答写入表格
        if not os.path.exists(file_name):
            wb = Workbook()
            wb.save(file_name)
            print(f"文件 {file_name} 已创建。")
        wb = load_workbook(file_name)
        ws = wb.active if wb.active else wb.create_sheet()
        max_row = ws.max_row
        ws[f'A{max_row + 1}'] = placeholder_text
        ws[f'B{max_row + 1}'] = text_answer
        ws[f'C{max_row + 1}'] = Accuracy

        wb.save(file_name)
        print("存入表格成功")

#点击侧边栏
def touch_expand():
    #xpath 判断侧边栏是否拉开 未拉开则需要拉开
    #此处xpath对应展开键
    expandkey_xpath_expression = '//*[@id="root"]/div/article/aside[1]/aside'

    #通过xpath寻找拉开键
    expandkey = driver.find_element(By.XPATH, expandkey_xpath_expression)

    assert_exist_key(expandkey)

#点击妙想助手
def touch_helper():
    #妙想助手按钮xpath
    helperkey_xpath_expression = '//*[@id="root"]/div/article/aside[1]/div/header/button[2]'

    #通过xpath寻找妙想助手按钮
    helperkey = driver.find_element(By.XPATH, helperkey_xpath_expression)

    assert_exist_key(helperkey)

#点击对应模块按钮
def touch_module(module_key):
    assert_exist_key(module_key)

#点击问题输入框
def touch_questionbox():
    #问题输入框
    questionbox_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/footer/main/textarea'

    #通过xpath寻找问题输入框
    questionbox = driver.find_element(By.XPATH, questionbox_xpath_expression)

    assert_exist_key(questionbox)

#自定义问题输入


#点击随机推荐问题
def touch_randomquestion():
    #推荐问题n按钮xpath
    n = random.randint(1,3)
    recommendquestionkey_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/main/div[1]/article/main/div/div/main/article/div['+str(n)+']'

    #通过xpath寻找推荐问题n按钮
    recommendquestionkey = driver.find_element(By.XPATH, recommendquestionkey_xpath_expression)

    assert_exist_key(recommendquestionkey)

#点击发送按键
def touch_send():
    #发送按钮xpath
    sendkey_xpath_expression = '//*[@id="MxChatRoomListWrap"]/div/div[1]/footer/main/section[2]'

    #通过xpath寻找发送按钮
    sendkey = driver.find_element(By.XPATH, sendkey_xpath_expression)

    assert_exist_key(sendkey)
        
    #等待回答加载完毕
    time.sleep(15)

#爬取输入框问题
def get_placeholder_text():
    global placeholder_text
    placeholder_text = driver.find_element(By.CLASS_NAME,"ChatRoom-module__input___BotWC")
    placeholder_text = placeholder_text.text
    placeholder_text = unescape(placeholder_text)
    #print(placeholder_text.text)
    #time.sleep(1)

#爬取前端文字回答
def get_text_answer():
    global text_answer
    text_answer = driver.find_element(By.CLASS_NAME,"ShowType40-module__content___CWmch")
    text_answer = text_answer.text
    text_answer = unescape(text_answer)
    #print(text_answer.text)
    #time.sleep(1)

#比对问题回答的相似度
def compare_similarity():
    numofwords_in_placeholder = 0
    numofwords_in_text_answer = 0
    words_in_placeholder = list(jieba.cut(placeholder_text,cut_all=False))
    words_in_text_answer = list(jieba.cut(text_answer,cut_all=False))
    numofwords_in_placeholder = sum(1 for _ in words_in_placeholder)
    #numofwords_in_text_answer = sum(1 for _ in words_in_text_answer)
    set_words_in_placeholder = set(words_in_placeholder)
    set_words_in_text_answer = set(words_in_text_answer)
    
    common_words = set_words_in_placeholder & set_words_in_text_answer

    Accuracy = float(len(common_words)) / float(numofwords_in_placeholder)
    assert_accuracy(Accuracy)


#进入妙想主页
driver.get("http://aife-test.eastmoney.com:9000/chat")
driver.maximize_window()
#先使用sleep 暴力等待
time.sleep(1)


#无限循环
while i < ROUND:

    touch_expand()
    touch_helper()

    #循环所选模块
    for module_xpath in module_xpath_expression:
        #通过xpath寻找模块按钮
        module_key = driver.find_element(By.XPATH, module_xpath)
        touch_module(module_key)

        touch_randomquestion()

        get_placeholder_text()
        
        touch_send()

        get_text_answer()

        compare_similarity()

        touch_expand()

        touch_helper()

    i = i + 1

driver.quit()
