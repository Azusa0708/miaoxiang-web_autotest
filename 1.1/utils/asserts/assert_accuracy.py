import os
from openpyxl import Workbook
from openpyxl import load_workbook
from config.config import ACCURACY,file_name
from utils.reptile.get_text import placeholder_text,text_answer

#结果精确度断言
def assert_accuracy(Accuracy):
    if Accuracy > ACCURACY:
        print("准确度合格：",Accuracy)
    else:
        print("准确度不合格，问答存入表格：",Accuracy)
        #问答写入表格
        if not os.path.exists(file_name):
            wb = Workbook()
            wb.save(file_name)
            print(f"文件 {file_name} 已创建。")
        wb = load_workbook(file_name)
        ws = wb.active if wb.active else wb.create_sheet()
        max_row = ws.max_row
        ws[f'A{max_row + 1}'] = placeholder_text
        ws[f'B{max_row + 1}'] = text_answer
        ws[f'C{max_row + 1}'] = Accuracy

        wb.save(file_name)
        print("存入表格成功")